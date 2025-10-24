"""
File processing utilities for data feeds.
Handles txt, csv, and xlsx file parsing and content extraction.
"""
from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 100 * 1024 * 1024  # 100 MB in bytes


def validate_file_size(file_size: int) -> tuple[bool, Optional[str]]:
    """
    Validate that file size is within acceptable limits.
    
    Args:
        file_size: Size of file in bytes
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if file_size > MAX_FILE_SIZE:
        size_mb = file_size / (1024 * 1024)
        return False, (
            f"File size ({size_mb:.2f} MB) exceeds the 100 MB limit. "
            "Please upload larger files to the designated SharePoint location."
        )
    return True, None


def process_txt(file_data: bytes, filename: str = "upload.txt") -> Dict[str, Any]:
    """
    Process a text file and extract content.
    
    Args:
        file_data: Raw file bytes
        filename: Original filename
        
    Returns:
        Dictionary with content, metadata, and processing info
    """
    try:
        # Try different encodings
        content = None
        encoding_used = None
        
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'ascii']:
            try:
                content = file_data.decode(encoding)
                encoding_used = encoding
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise ValueError("Unable to decode file with supported encodings")
        
        # Extract metadata
        line_count = len(content.splitlines())
        word_count = len(content.split())
        char_count = len(content)
        
        return {
            "content": content,
            "processed_content": content.strip(),
            "metadata": {
                "filename": filename,
                "encoding": encoding_used,
                "line_count": line_count,
                "word_count": word_count,
                "char_count": char_count,
                "file_type": "txt",
            },
            "success": True,
        }
        
    except Exception as exc:
        logger.error(f"Error processing txt file: {exc}", exc_info=True)
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": str(exc), "file_type": "txt"},
            "success": False,
        }


def process_csv(file_data: bytes, filename: str = "upload.csv") -> Dict[str, Any]:
    """
    Process a CSV file and extract structured content.
    
    Args:
        file_data: Raw file bytes
        filename: Original filename
        
    Returns:
        Dictionary with content, metadata, row count, and columns
    """
    try:
        # Decode content
        content = None
        encoding_used = None
        
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                content = file_data.decode(encoding)
                encoding_used = encoding
                break
            except UnicodeDecodeError:
                continue
        
        if content is None:
            raise ValueError("Unable to decode CSV file")
        
        # Parse CSV
        csv_reader = csv.reader(io.StringIO(content))
        rows = list(csv_reader)
        
        if not rows:
            raise ValueError("CSV file is empty")
        
        # Extract headers and data
        headers = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        
        # Create text representation for embedding
        text_parts = []
        text_parts.append(f"CSV file with columns: {', '.join(headers)}")
        
        # Add sample rows (first 5 for context)
        sample_size = min(5, len(data_rows))
        for i, row in enumerate(data_rows[:sample_size]):
            row_text = " | ".join(str(cell) for cell in row)
            text_parts.append(f"Row {i+1}: {row_text}")
        
        if len(data_rows) > sample_size:
            text_parts.append(f"... and {len(data_rows) - sample_size} more rows")
        
        processed_content = "\n".join(text_parts)
        
        return {
            "content": content,
            "processed_content": processed_content,
            "metadata": {
                "filename": filename,
                "encoding": encoding_used,
                "row_count": len(data_rows),
                "total_rows": len(rows),
                "column_count": len(headers),
                "columns": headers,
                "file_type": "csv",
            },
            "rows": data_rows,
            "columns": headers,
            "success": True,
        }
        
    except Exception as exc:
        logger.error(f"Error processing CSV file: {exc}", exc_info=True)
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": str(exc), "file_type": "csv"},
            "success": False,
        }


def process_xlsx(file_data: bytes, filename: str = "upload.xlsx") -> Dict[str, Any]:
    """
    Process an Excel file and extract structured content.
    
    Args:
        file_data: Raw file bytes
        filename: Original filename
        
    Returns:
        Dictionary with content, metadata, row count, and columns
    """
    try:
        # Try to import openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.warning("openpyxl not installed, cannot process xlsx files")
            return {
                "content": "",
                "processed_content": "",
                "metadata": {
                    "error": "Excel processing requires openpyxl package",
                    "file_type": "xlsx",
                },
                "success": False,
            }
        
        # Load workbook from bytes
        workbook = load_workbook(filename=io.BytesIO(file_data), read_only=True, data_only=True)

        sheets_data: List[Dict[str, Any]] = []
        overall_text_parts: List[str] = []

        for sheet in workbook.worksheets:
            rows: List[List[str]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(cell) if cell is not None else "" for cell in row])

            if not rows:
                headers: List[str] = []
                data_rows: List[List[str]] = []
            else:
                headers = rows[0]
                data_rows = rows[1:] if len(rows) > 1 else []

            # Per-sheet summary
            overall_text_parts.append(
                f"Excel sheet '{sheet.title}' with columns: {', '.join(headers)}"
            )
            sample_size = min(5, len(data_rows))
            for i, row in enumerate(data_rows[:sample_size]):
                row_text = " | ".join(str(cell) for cell in row)
                overall_text_parts.append(f"{sheet.title} Row {i+1}: {row_text}")
            if len(data_rows) > sample_size:
                overall_text_parts.append(
                    f"{sheet.title}: ... and {len(data_rows) - sample_size} more rows"
                )

            sheets_data.append(
                {
                    "sheet_name": sheet.title,
                    "rows": data_rows,
                    "columns": headers,
                    "row_count": len(data_rows),
                    "column_count": len(headers),
                }
            )

        processed_content = "\n".join(overall_text_parts) if overall_text_parts else "Excel file with no readable sheets"

        # Provide a CSV-like string for the first sheet (back-compat)
        content = ""
        if sheets_data:
            csv_content = io.StringIO()
            writer = csv.writer(csv_content)
            first_sheet = sheets_data[0]
            if first_sheet["columns"]:
                writer.writerow(first_sheet["columns"])
            for row in first_sheet["rows"]:
                writer.writerow(row)
            content = csv_content.getvalue()

        workbook.close()

        # Back-compat top-level rows/columns from first sheet
        top_rows = sheets_data[0]["rows"] if sheets_data else []
        top_cols = sheets_data[0]["columns"] if sheets_data else []

        return {
            "content": content,
            "processed_content": processed_content,
            "metadata": {
                "filename": filename,
                "file_type": "xlsx",
                "sheets_count": len(sheets_data),
                "sheet_names": [s["sheet_name"] for s in sheets_data],
                "row_count": sum(s["row_count"] for s in sheets_data) if sheets_data else 0,
                "column_count": len(top_cols),
                "columns": top_cols,
            },
            "rows": top_rows,
            "columns": top_cols,
            "sheets": sheets_data,
            "success": True,
        }
        
    except Exception as exc:
        logger.error(f"Error processing Excel file: {exc}", exc_info=True)
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": str(exc), "file_type": "xlsx"},
            "success": False,
        }


def process_json(file_data: bytes, filename: str = "upload.json") -> Dict[str, Any]:
    """Process a JSON file and extract structured content.

    - Supports objects, arrays of objects, nested structures
    - Returns a summary text, inferred schema, and optional flattened tabular view for simple arrays
    """
    try:
        # Decode using utf-8 and fallback
        content_str: Optional[str] = None
        for enc in ["utf-8", "cp1252", "latin-1"]:
            try:
                content_str = file_data.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if content_str is None:
            raise ValueError("Unable to decode JSON file")

        data = json.loads(content_str)

        def infer_schema(obj) -> Any:
            if isinstance(obj, dict):
                return {k: infer_schema(v) for k, v in obj.items()}
            if isinstance(obj, list):
                # Schema of list is schema of first non-null element
                for el in obj:
                    if el is not None:
                        return [infer_schema(el)]
                return []
            return type(obj).__name__

        schema = infer_schema(data)

        # Try to flatten simple array of dicts for tabular metrics
        rows: List[List[Any]] = []
        columns: List[str] = []
        row_count = 0
        if isinstance(data, list) and data and all(isinstance(x, dict) for x in data):
            # Collect union of keys
            keys: List[str] = sorted({k for d in data for k in d.keys()})
            columns = keys
            for d in data:
                rows.append([d.get(k, None) for k in keys])
            row_count = len(rows)

        # Build summary text
        text_parts: List[str] = [
            f"JSON file '{filename}'",
            f"Top-level type: {type(data).__name__}",
        ]
        if isinstance(data, dict):
            text_parts.append(f"Top-level keys: {', '.join(list(data.keys())[:30])}")
        if columns:
            text_parts.append(f"Tabular view columns: {', '.join(columns[:30])}")
            sample_size = min(5, row_count)
            for i in range(sample_size):
                text_parts.append("Row " + str(i + 1) + ": " + " | ".join(str(v) for v in rows[i]))
            if row_count > sample_size:
                text_parts.append(f"... and {row_count - sample_size} more rows")

        processed_content = "\n".join(text_parts)

        return {
            "content": content_str,
            "processed_content": processed_content,
            "metadata": {
                "filename": filename,
                "file_type": "json",
                "row_count": row_count,
                "column_count": len(columns),
                "columns": columns,
                "schema": schema,
            },
            "rows": rows,
            "columns": columns,
            "json_data": data,
            "success": True,
        }
    except Exception as exc:
        logger.error(f"Error processing JSON file: {exc}", exc_info=True)
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": str(exc), "file_type": "json"},
            "success": False,
        }


def process_file(file_data: bytes, filename: str, file_type: str) -> Dict[str, Any]:
    """
    Process a file based on its type.
    
    Args:
        file_data: Raw file bytes
        filename: Original filename
        file_type: Type of file (txt, csv, xlsx)
        
    Returns:
        Processing result dictionary
    """
    # Validate file size
    is_valid, error_msg = validate_file_size(len(file_data))
    if not is_valid:
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": error_msg, "file_type": file_type},
            "success": False,
        }
    
    # Process based on type
    file_type = file_type.lower()
    
    if file_type == "txt":
        return process_txt(file_data, filename)
    elif file_type == "csv":
        return process_csv(file_data, filename)
    elif file_type == "xlsx":
        return process_xlsx(file_data, filename)
    elif file_type == "json":
        return process_json(file_data, filename)
    else:
        return {
            "content": "",
            "processed_content": "",
            "metadata": {
                "error": f"Unsupported file type: {file_type}",
                "file_type": file_type,
            },
            "success": False,
        }


def process_text_input(text: str, name: str = "Text Input") -> Dict[str, Any]:
    """
    Process direct text input.
    
    Args:
        text: Input text
        name: Name/title for the input
        
    Returns:
        Processing result dictionary
    """
    try:
        if not text or not text.strip():
            return {
                "content": "",
                "processed_content": "",
                "metadata": {"error": "Text input is empty", "file_type": "text_input"},
                "success": False,
            }
        
        # Calculate metadata
        line_count = len(text.splitlines())
        word_count = len(text.split())
        char_count = len(text)
        
        return {
            "content": text,
            "processed_content": text.strip(),
            "metadata": {
                "name": name,
                "line_count": line_count,
                "word_count": word_count,
                "char_count": char_count,
                "file_type": "text_input",
            },
            "success": True,
        }
        
    except Exception as exc:
        logger.error(f"Error processing text input: {exc}", exc_info=True)
        return {
            "content": "",
            "processed_content": "",
            "metadata": {"error": str(exc), "file_type": "text_input"},
            "success": False,
        }

